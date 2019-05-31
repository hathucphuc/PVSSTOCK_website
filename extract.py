from openpyxl import load_workbook

Input_Xlsx_FileName = "/Users/haphuc/Downloads/testUnmerge.xlsx"
Output_Xlsx_FileName = "/Users/haphuc/Downloads/OutputtestUnmerge.xlsx"

Book = load_workbook(Input_Xlsx_FileName)
Worksheet = Book['Hoja1']

Worksheet.unmerge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
# Worksheet.merge_cells("A3:B3")
# Worksheet.unmerge_cells("A3:B3")
Worksheet.cell(1,1).value = "TOTO"
Worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
# Worksheet.merge_cells("A3:B3")

Book.save(filename = Output_Xlsx_FileName)